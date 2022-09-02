import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart      # Многокомпонентный объект

with open("email_template.html") as file:
			html = file.read()


book = openpyxl.open('sender_base.xlsx')
sheet = book.active

for i in range(1, sheet.max_row+1):
    # desk_email = sheet[i][0].value
    desk_email = sheet.cell(row=i, column=1).value
    print(desk_email)
# for e in range(1, sheet.max_row+1):
#     # desk_email = sheet[i][0].value
    sheet.delete_cols(row=i, column=1)
# sheet.cell(row=2, column=3).value = "Ok"
    # sheet.append(row=2, column=2).value = 'Biggest Indian Festival for Hindus'
    # file.save ('sender_base.xlsx')



# sheet.cell(row=i, column=1).value